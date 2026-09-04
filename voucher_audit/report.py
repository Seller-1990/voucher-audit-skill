from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import yaml

from .report_customer_consistency import build_customer_consistency_sheet
from .report_pp_change import build_pp_change_sheet
from .report_comparison import build_comparison_report
from .report_cost_checks import build_outsourcing_missing_cost_sheet, build_rev_cost_zero_mismatch_sheet
from .report_profit import build_neg_profit_ratio_sheet
from .deep_analysis import (
    build_correlation_index,
    build_customer_profile_sheet,
    build_fix_list_sheet,
    build_rule_correlation_sheet,
)
from .registration import (
    attach_registration_status,
    export_registration_rows,
    find_registration_file,
    load_registration_table,
)

# 数值精度：金额 2 位小数；比率/占比类指标 4 位小数
_RATIO_HINTS = ("率", "/", "占比", "ratio", "％", "%")


def _round_for_output(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    indicator_names: dict[int, str] = {}
    if "指标名称" in out.columns:
        indicator_names = {i: str(v) for i, v in out["指标名称"].items()}

    def _round_value(i: int, v: object, default_nd: int) -> object:
        try:
            f = float(v)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return v
        if pd.isna(f):
            return v
        name = indicator_names.get(i, "")
        nd = 4 if any(h in name.lower() for h in _RATIO_HINTS) else default_nd
        return round(f, nd)

    for col in out.columns:
        if col == "指标值":
            # 指标值列常为 object 混合类型，逐格转换取整
            out[col] = [_round_value(i, v, 2) for i, v in enumerate(out[col])]
        elif pd.api.types.is_float_dtype(out[col]):
            out[col] = out[col].round(2)
    return out


def _sort_by_mark(df: pd.DataFrame) -> pd.DataFrame:
    """有“标注”列的 sheet：问题行置顶、参考行靠后（组内保持原序）。"""
    if df is None or df.empty or "标注" not in df.columns:
        return df
    rank = {"问题": 0, "错误": 0, "需确认": 0, "参考": 1}
    key = df["标注"].map(rank).fillna(2)
    return df.assign(__mark_rank=key).sort_index(kind="stable").sort_values(
        by="__mark_rank", kind="stable"
    ).drop(columns=["__mark_rank"])


# 新增规则（收入成本表1.py + 登记表 整合）明细页规则ID集合
_EXTRA_RULE_IDS = {
    "INC_MOM_CHANGE",
    "INC_GM_HIGH_RATIO",
    "INC_REV_COST_INVERSION",
    "INC_HEADCOUNT_REV_MISMATCH",
    "INC_SOCIAL_HEADCOUNT_MISMATCH",
    "INC_COST_RATIO_HIGH",
    "INC_EXPENSE_RATIO",
    "INC_COST_SUDDEN_APPEARANCE",
    "INC_DUPLICATE_ROW",
    "INC_GROUP_HQ_UNSETTLED",
    "INC_SIMILAR_CUSTOMER_RENAME",
    "AUX_WAGE_WRONG_CUSTOMER",
    "INC_MIXED_BIZ_TYPE",
}


def build_extra_rules_sheet(
    income_dim_anomalies: Optional[pd.DataFrame],
    income_gm_anomalies: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """整合规则明细页：展示新增规则（高毛利率/倒挂/人次社保背离/占比/环比等）的命中明细。"""
    parts: list[pd.DataFrame] = []
    for df in (income_dim_anomalies, income_gm_anomalies):
        if df is None or df.empty or "规则ID" not in df.columns:
            continue
        sub = df[df["规则ID"].astype(str).isin(_EXTRA_RULE_IDS)].copy()
        if not sub.empty:
            parts.append(sub)
    if not parts:
        return pd.DataFrame()
    out = pd.concat(parts, ignore_index=True)
    out = _strip_internal_columns(out)
    out = _replace_rule_id_with_name(out)
    keep = [c for c in ["严重度", "规则名称", "规则ID", "主体账簿", "月", "三级科目", "实际客户", "部门", "项目", "命中原因", "指标值"] if c in out.columns]
    rest = [c for c in out.columns if c not in keep and c != "命中原因"]
    order = keep + [c for c in out.columns if c not in keep]
    out = out[order]
    if "严重度" in out.columns:
        out = out.assign(__sev=out["严重度"].map(_severity_rank)).sort_values(by=["__sev"]).drop(columns=["__sev"])
    return out.reset_index(drop=True)


def _write_sheet(w: pd.ExcelWriter, name: str, df: pd.DataFrame) -> None:
    ws_name = str(name)[:31]
    df.to_excel(w, sheet_name=ws_name, index=False)
    try:
        from openpyxl.styles import Font, PatternFill

        ws = w.sheets[ws_name]
        if ws.max_row > 1 and ws.max_column > 0:
            from openpyxl.utils import get_column_letter

            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        # 样式：冻结首行 + 按内容自适应列宽
        ws.freeze_panes = "A2"
        headers = [str(c.value) for c in ws[1]]
        # 疑似数据错误清单：优先级底纹（P1红/P2黄/P3灰，参考 收入成本表1.py 标注风格）
        if "优先级" in headers:
            pri_idx = headers.index("优先级") + 1
            fills = {
                "P1 疑似错误": PatternFill("solid", fgColor="FFC7CE"),
                "P2 需确认": PatternFill("solid", fgColor="FFF2CC"),
                "P3 波动参考": PatternFill("solid", fgColor="D9D9D9"),
            }
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=pri_idx).value
                if v in fills:
                    ws.cell(row=r, column=pri_idx).fill = fills[v]
        # 含“毛利率”列的页：毛利率<0 红字、>15% 绿字、其余百分比格式（同 收入成本表1.py 毛利异常项目）
        if "毛利率" in headers:
            gm_idx = headers.index("毛利率") + 1
            red_fill = PatternFill("solid", fgColor="FFC7CE")
            red_font = Font(color="9C0006")
            green_fill = PatternFill("solid", fgColor="C6EFCE")
            green_font = Font(color="006100")
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=gm_idx)
                v = c.value
                if isinstance(v, (int, float)):
                    c.number_format = "0.00%"
                    if v < 0:
                        c.fill = red_fill
                        c.font = red_font
                    elif v > 0.15:
                        c.fill = green_fill
                        c.font = green_font
        # 毛利类页（负毛利占比检查）：项目毛利润<0 红标（同 收入成本表1.py 项目毛利为负项目）
        if "毛利" in ws_name and "项目毛利润" in headers:
            profit_idx = headers.index("项目毛利润") + 1
            red_fill = PatternFill("solid", fgColor="FFC7CE")
            red_font = Font(color="9C0006")
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=profit_idx)
                v = c.value
                if isinstance(v, (int, float)) and v < 0:
                    c.fill = red_fill
                    c.font = red_font
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col[: min(len(col), 200)])
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 45)
    except Exception:
        pass


def _severity_rank(s: str) -> int:
    s = str(s)
    if s == "错误":
        return 0
    if s == "需确认":
        return 1
    return 2


@dataclass(frozen=True)
class ReportPaths:
    output_dir: Path
    report_path: Path


@dataclass(frozen=True)
class RuleInfo:
    """规则信息数据类"""
    rule_id: str
    rule_name: str
    severity: str
    scope: str
    description: str
    source_doc: str
    source_clause: str
    params_summary: str = ""


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def make_report_paths(workdir: Path, report_prefix: str, yyyymm: str) -> ReportPaths:
    out_dir = workdir / "凭证审核输出"
    ensure_dir(out_dir)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    report_path = out_dir / f"{report_prefix}_{yyyymm}_{ts}.xlsx"
    return ReportPaths(output_dir=out_dir, report_path=report_path)


def _safe_text(v: object) -> str:
    if v is None:
        return "空"
    if isinstance(v, float) and pd.isna(v):
        return "空"
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return "空"
    return s


def _sheet_name(report_format: dict[str, Any], logical_key: str, default_name: str) -> str:
    sheet_names = report_format.get("sheet_names", {}) if isinstance(report_format, dict) else {}
    name = default_name
    if isinstance(sheet_names, dict):
        cand = sheet_names.get(logical_key)
        if isinstance(cand, str) and cand.strip():
            name = cand.strip()
    return name[:31]


def _strip_internal_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    keep = [c for c in df.columns if not str(c).startswith("_")]
    return df[keep].copy() if keep else pd.DataFrame()


def _replace_rule_id_with_name(df: pd.DataFrame) -> pd.DataFrame:
    """Replace '规则ID' column values with '规则名称' values, then drop '规则ID' column."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "规则ID" in out.columns and "规则名称" in out.columns:
        out["规则ID"] = out["规则名称"]
    if "规则名称" in out.columns:
        out = out.drop(columns=["规则名称"])
    if "规则ID" in out.columns:
        out = out.rename(columns={"规则ID": "规则名称"})
    return out


def _apply_column_layout(df: pd.DataFrame, report_format: dict[str, Any], logical_key: str) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    df = _strip_internal_columns(df)
    layouts = report_format.get("column_layouts", {}) if isinstance(report_format, dict) else {}
    if not isinstance(layouts, dict):
        return df
    layout = layouts.get(logical_key, {})
    if not isinstance(layout, dict):
        return df

    out = df.copy()
    rename_map = layout.get("rename", {})
    if isinstance(rename_map, dict):
        clean_rename = {str(k): str(v) for k, v in rename_map.items() if str(k) in out.columns and str(v).strip()}
        if clean_rename:
            out = out.rename(columns=clean_rename)

    keep_cols = layout.get("keep", None)
    if isinstance(keep_cols, list):
        keep = [str(x) for x in keep_cols if str(x) in out.columns]
        if keep:
            out = out[keep]

    order = layout.get("order", None)
    if isinstance(order, list):
        ordered = [str(x) for x in order if str(x) in out.columns]
        rest = [c for c in out.columns if c not in ordered]
        if ordered:
            out = out[ordered + rest]
    return out


def _build_combo_drift_friendly_view(income_dim_anomalies: pd.DataFrame) -> pd.DataFrame:
    if income_dim_anomalies is None or income_dim_anomalies.empty:
        return pd.DataFrame()
    if "规则ID" not in income_dim_anomalies.columns:
        return pd.DataFrame()

    df = income_dim_anomalies[income_dim_anomalies["规则ID"].astype(str) == "INC_CUSTOMER_CONSISTENCY"].copy()
    if df.empty:
        return pd.DataFrame()

    priorities = ["三级科目", "实际客户", "部门", "项目"]

    def build_issue(row: pd.Series) -> pd.Series:
        issue_points: list[str] = []
        history_items: list[str] = []
        primary = ""
        for f in priorities:
            hist_col = f"历史主_{f}"
            cur_col = f"本月主_{f}"
            if hist_col not in row or cur_col not in row:
                continue
            if _safe_text(row.get(hist_col)) != _safe_text(row.get(cur_col)):
                tag = f"{f}疑似异常"
                issue_points.append(tag)
                history_items.append(f"{f}历史={_safe_text(row.get(hist_col))}")
                if not primary:
                    primary = tag
        return pd.Series({"主问题分类": primary, "问题点": "；".join(issue_points), "历史对应信息": "；".join(history_items)})

    if not {"主问题分类", "问题点", "历史对应信息"}.issubset(df.columns):
        df = pd.concat([df, df.apply(build_issue, axis=1)], axis=1)
    else:
        df["主问题分类"] = df["主问题分类"].astype(str)
        df["问题点"] = df["问题点"].astype(str)
        df["历史对应信息"] = df["历史对应信息"].astype(str)
        missing_mask = df["问题点"].str.strip().isin(["", "nan"])
        if missing_mask.any():
            filled = df.loc[missing_mask].apply(build_issue, axis=1)
            df.loc[missing_mask, "主问题分类"] = filled["主问题分类"]
            df.loc[missing_mask, "问题点"] = filled["问题点"]
            df.loc[missing_mask, "历史对应信息"] = filled["历史对应信息"]

    if "对比月份" not in df.columns:
        if {"本期月份", "前期月份"}.issubset(df.columns):
            df["对比月份"] = df["本期月份"].map(lambda x: _safe_text(x)) + " vs " + df["前期月份"].map(lambda x: _safe_text(x))
        else:
            df["对比月份"] = ""

    if "影响金额" not in df.columns:
        if "cur_total_abs" in df.columns:
            df["影响金额"] = pd.to_numeric(df["cur_total_abs"], errors="coerce").fillna(0.0)
        else:
            df["影响金额"] = 0.0

    rename_map = {
        "本月主_三级科目": "本月三级科目",
        "本月主_实际客户": "本月实际客户",
        "本月主_部门": "本月部门",
        "本月主_项目": "本月项目",
    }
    df = df.rename(columns=rename_map)

    cols = [
        "严重度",
        "规则ID",
        "主问题分类",
        "问题点",
        "历史对应信息",
        "主体账簿",
        "账载客户",
        "本月三级科目",
        "本月实际客户",
        "本月部门",
        "本月项目",
        "对比月份",
        "影响金额",
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    out = df[cols].copy()
    if "影响金额" in out.columns:
        out["影响金额"] = pd.to_numeric(out["影响金额"], errors="coerce").fillna(0.0)
        out = out.sort_values(by=["影响金额"], ascending=False)
    return out


def write_report(
    path: Path,
    overview: pd.DataFrame,
    overview_rule_breakdown: Optional[pd.DataFrame] = None,
    aux_rule_violations: pd.DataFrame = None,
    aux_suspect_wrong_account: pd.DataFrame = None,
    income_dim_anomalies: pd.DataFrame = None,
    income_gm_anomalies: pd.DataFrame = None,
    ai_review: Optional[pd.DataFrame] = None,
    report_format: Optional[dict[str, Any]] = None,
    # 新增参数
    checks: Optional[list[dict[str, Any]]] = None,
    df_income: Optional[pd.DataFrame] = None,
    df_aux: Optional[pd.DataFrame] = None,
    df_mapping: Optional[pd.DataFrame] = None,
    target_month: Optional[int] = None,
    workdir: Optional[Path] = None,
    yyyymm: str = "",
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        # 1. 规则说明
        rule_info_out = build_rule_info_sheet(checks or [])
        if rule_info_out is None or rule_info_out.empty:
            rule_info_out = pd.DataFrame({"提示": ["未提供规则清单"]})
        _write_sheet(w, "规则与内容", _round_for_output(rule_info_out))

        # 2. 审核汇总（工作目录/目标月/各类命中总数/AI 状态等）
        overview_out = overview if overview is not None and not overview.empty else pd.DataFrame({"提示": ["无汇总信息"]})
        _write_sheet(w, "审核汇总", _round_for_output(overview_out))

        # 3. 规则命中统计（每规则命中数与严重度分布）
        breakdown_out = (
            overview_rule_breakdown
            if overview_rule_breakdown is not None and not overview_rule_breakdown.empty
            else pd.DataFrame({"提示": ["无命中"]})
        )
        _write_sheet(w, "规则命中统计", _round_for_output(breakdown_out))

        # 4. 疑似数据错误清单（面向当期修正的统一输出：P1错误/P2需确认/P3参考 + 修正动作）
        fix_out = build_fix_list_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            income_gm_anomalies=income_gm_anomalies,
            aux_rule_violations=aux_rule_violations,
            aux_df=df_aux,
            target_month=target_month,
        )
        # 登记表打通：已知问题去重（登记状态列）+ 待登记行导出
        reg_path = find_registration_file(workdir) if workdir is not None else None
        reg_df = load_registration_table(reg_path)
        if not reg_df.empty and fix_out is not None and not fix_out.empty:
            fix_out = attach_registration_status(fix_out, reg_df, yyyymm)
        elif fix_out is not None and not fix_out.empty:
            fix_out = fix_out.copy()
            fix_out["登记状态"] = "未登记"
        if fix_out is None or fix_out.empty:
            fix_out = pd.DataFrame({"提示": ["无疑似数据错误命中"]})
        _write_sheet(w, "疑似数据错误清单", fix_out)

        # 4.1 待登记行（登记表格式，可直接粘贴走指派闭环）
        reg_export = export_registration_rows(fix_out, yyyymm)
        if reg_export is None or reg_export.empty:
            reg_export = pd.DataFrame({"提示": ["无待登记项（全部已登记或无命中）"]})
        _write_sheet(w, "待登记异常", reg_export)

        # 4.1 新增规则明细（高毛利率/倒挂/人次社保背离/占比/环比等）
        extra_out = build_extra_rules_sheet(income_dim_anomalies, income_gm_anomalies)
        if extra_out is None or extra_out.empty:
            extra_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "新增规则明细", _round_for_output(extra_out))

        # 5. 深度分析：规则关联影响 + 客户综合分析
        correlation_index, pp_keys = build_correlation_index(income_dim_anomalies, income_gm_anomalies)
        corr_out = build_rule_correlation_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            income_gm_anomalies=income_gm_anomalies,
            aux_rule_violations=aux_rule_violations,
            target_month=target_month,
        )
        corr_out = _sort_by_mark(corr_out)
        if corr_out is None or corr_out.empty:
            corr_out = pd.DataFrame({"提示": ["无多规则关联命中"]})
        _write_sheet(w, "规则关联分析", corr_out)

        profile_out = build_customer_profile_sheet(
            df_income=df_income,
            target_month=target_month,
            correlation_index=correlation_index,
            pp_keys=pp_keys,
        )
        if profile_out is None or profile_out.empty:
            profile_out = pd.DataFrame({"提示": ["无客户画像数据"]})
        _write_sheet(w, "客户综合分析", profile_out)

        # 6. 明细 sheets
        headcount_out = build_headcount_report(
            df_aux=df_aux,
            aux_rule_violations=aux_rule_violations,
            aux_suspect_wrong_account=aux_suspect_wrong_account,
        )
        if headcount_out is None or headcount_out.empty:
            headcount_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "人次数据检查", _round_for_output(headcount_out))

        customer_out = build_customer_consistency_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            df_mapping=df_mapping,
            target_month=target_month,
        )
        customer_out = _sort_by_mark(customer_out)
        if customer_out is None or customer_out.empty:
            customer_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "客户归属一致性检查", _round_for_output(customer_out))

        rev_cost_out = build_rev_cost_zero_mismatch_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            checks=checks,
            target_month=target_month,
        )
        if rev_cost_out is None or rev_cost_out.empty:
            rev_cost_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "收入成本零值不匹配检查", _round_for_output(rev_cost_out))

        pp_out = build_pp_change_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            checks=checks,
            target_month=target_month,
        )
        pp_out = _sort_by_mark(pp_out)
        if pp_out is None or pp_out.empty:
            pp_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "同比波动检查", _round_for_output(pp_out))

        outsourcing_out = build_outsourcing_missing_cost_sheet(
            df_income=df_income,
            income_dim_anomalies=income_dim_anomalies,
            checks=checks,
            target_month=target_month,
        )
        if outsourcing_out is None or outsourcing_out.empty:
            outsourcing_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "外包缺工资或挂靠检查", _round_for_output(outsourcing_out))

        neg_gm_out = build_neg_profit_ratio_sheet(
            df_income=df_income,
            income_gm_anomalies=income_gm_anomalies,
            checks=checks,
            target_month=target_month,
        )
        if neg_gm_out is None or neg_gm_out.empty:
            neg_gm_out = pd.DataFrame({"提示": ["无命中"]})
        _write_sheet(w, "负毛利占比检查", _round_for_output(neg_gm_out))


def build_comparison_report_for_all_rules(
    df_income: pd.DataFrame,
    income_dim_anomalies: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """
    为所有有对比数据的规则生成专门报告
    当前只处理客户归属一致性检查（已完善），其他规则待后续迭代处理。

    返回: {sheet_name: report_df}
    """
    if income_dim_anomalies is None or income_dim_anomalies.empty:
        return {}

    if "规则ID" not in income_dim_anomalies.columns:
        return {}

    # 当前只保留已完善的规则：客户归属一致性检查
    comparison_rules = {
        "INC_CUSTOMER_CONSISTENCY": "客户归属一致性检查",
    }

    reports: dict[str, pd.DataFrame] = {}

    for rule_id, sheet_name in comparison_rules.items():
        rule_hits = income_dim_anomalies[income_dim_anomalies["规则ID"] == rule_id].copy()
        if not rule_hits.empty:
            report = build_comparison_report(
                df_source=df_income,
                hits_df=rule_hits,
                rule_id=rule_id,
            )
            if not report.empty:
                reports[sheet_name] = report

    return reports


def build_headcount_report(
    df_aux: Optional[pd.DataFrame],
    aux_rule_violations: Optional[pd.DataFrame],
    aux_suspect_wrong_account: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """
    生成人次数据检查专门报告（简化格式）

    格式：
    1. 只显示原记录的关键列（凭证号、摘要、命中码、严重度、命中原因）
    2. 过滤掉内部字段
    """
    frames: list[pd.DataFrame] = []
    if aux_rule_violations is not None and not aux_rule_violations.empty:
        frames.append(aux_rule_violations)
    if aux_suspect_wrong_account is not None and not aux_suspect_wrong_account.empty:
        frames.append(aux_suspect_wrong_account)
    if not frames:
        return pd.DataFrame()

    all_hits = pd.concat(frames, ignore_index=True)
    if all_hits.empty:
        return pd.DataFrame()

    if "规则ID" in all_hits.columns:
        mask = all_hits["规则ID"].astype(str).str.contains("AUX_HEADCOUNT", na=False)
    elif "规则名称" in all_hits.columns:
        mask = all_hits["规则名称"].astype(str).str.contains("人次", na=False)
    else:
        mask = pd.Series(False, index=all_hits.index)
    hits = all_hits[mask].copy()
    if hits.empty:
        return pd.DataFrame()

    keep_hit_cols = [c for c in ["命中码", "问题分类", "命中原因", "严重度"] if c in hits.columns]
    if df_aux is None or df_aux.empty or "_row_index" not in hits.columns:
        out = hits[keep_hit_cols].copy() if keep_hit_cols else hits.copy()
        drop_cols = [c for c in out.columns if c in ["规则ID", "规则名称", "规则描述", "制度来源"]]
        if drop_cols:
            out = out.drop(columns=drop_cols)
        return out

    idx_list = pd.to_numeric(hits["_row_index"], errors="coerce").fillna(-1).astype(int).tolist()
    detail = df_aux.reindex(idx_list).reset_index(drop=True)
    detail_cols = [
        c
        for c in detail.columns
        if not str(c).startswith("_")
        and not str(c).startswith("Unnamed:")
        and "凭证审核" not in str(c)
    ]
    out = detail[detail_cols].copy() if detail_cols else detail.copy()
    out["源行号"] = [int(i) + 2 if i >= 0 else "" for i in idx_list]

    for c in keep_hit_cols:
        out[c] = hits[c].reset_index(drop=True)

    final_cols = detail_cols + keep_hit_cols
    for c in final_cols:
        if c not in out.columns:
            out[c] = ""
    return out[final_cols]


def build_rule_info_sheet(checks: list[dict[str, Any]]) -> pd.DataFrame:
    """生成规则说明 sheet"""
    rows: list[dict[str, Any]] = []

    def sheet_for_rule(rule: dict[str, Any]) -> str:
        rtype = str(rule.get("type", "")).strip()
        if rtype == "customer_consistency_check":
            return "客户归属一致性检查"
        if rtype == "headcount_data_check":
            return "人次数据检查"
        if rtype == "pp_change":
            return "同比波动检查"
        if rtype == "neg_profit_ratio":
            return "负毛利占比检查"
        return "其他"

    sheet_order = ["人次数据检查", "客户归属一致性检查", "同比波动检查", "负毛利占比检查"]
    sheet_rank = {name: i for i, name in enumerate(sheet_order)}

    for i, rule in enumerate(checks):
        if not rule:
            continue
        source = rule.get("source", {}) or {}
        params = rule.get("params", {}) or {}

        sheet_name = sheet_for_rule(rule)
        rank = int(sheet_rank.get(sheet_name, 999))

        params_yaml = ""
        if isinstance(params, dict):
            try:
                params_yaml = yaml.safe_dump(params, allow_unicode=True, sort_keys=False).strip()
            except Exception:
                params_yaml = str(params)
        else:
            params_yaml = str(params)

        # 提取关键参数摘要
        params_summary = ""
        if isinstance(params, dict):
            key_params = []
            for k, v in params.items():
                if k in ("min_amount_abs", "tolerance_ratio", "min_gross_revenue", "threshold"):
                    key_params.append(f"{k}={v}")
            params_summary = ", ".join(key_params)

        rows.append({
            "__sheet_rank": rank,
            "__rule_index": i,
            "所属Sheet": sheet_name,
            "规则ID": str(rule.get("id", "")),
            "规则名称": str(rule.get("name", "")),
            "规则类型": str(rule.get("type", "")),
            "严重度": str(rule.get("severity", "")),
            "数据范围": str(rule.get("scope", "")),
            "规则描述": str(rule.get("description", "")),
            "制度来源": str(source.get("doc", "")),
            "制度条款": str(source.get("clause", "")),
            "关键参数": params_summary,
            "规则内容": params_yaml,
        })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out = out.sort_values(by=["__sheet_rank", "__rule_index"], ascending=True)
    out = out.drop(columns=["__sheet_rank", "__rule_index"], errors="ignore")
    return out


def _get_original_income_columns(df_income: pd.DataFrame) -> list[str]:
    """获取收入成本表的原始列名（用于保持原样输出）"""
    if df_income is None or df_income.empty:
        return []
    return [c for c in df_income.columns if not str(c).startswith("_")]


def build_customer_consistency_report(
    df_income: pd.DataFrame,
    customer_hits: pd.DataFrame,
    df_mapping: pd.DataFrame,
    target_month: int,
) -> pd.DataFrame:
    """
    生成客户归属一致性检查报告（简化格式）

    格式要求：
    1. 显示原记录（原始数据中的列，直接复制）
    2. 在原记录后增加"命中原因"列
    3. 对于对比类问题，在问题记录下方插入对比记录
    4. 命中原因可以有多个（不超过3条）
    """
    if customer_hits is None or customer_hits.empty:
        return pd.DataFrame()

    # 获取原始列名
    original_cols = _get_original_income_columns(df_income)

    all_rows: list[dict[str, Any]] = []

    for _, row in customer_hits.iterrows():
        # 基础命中信息
        hit_reasons: list[str] = []
        if "命中原因" in row and pd.notna(row["命中原因"]):
            hit_reasons.append(str(row["命中原因"]))

        # 构建问题记录行
        record: dict[str, Any] = {"_record_type": "问题"}

        # ① 原记录列（直接从原始数据复制）
        for col in original_cols:
            record[col] = row.get(col, "")

        # ② 命中原因列
        record["命中原因"] = "；".join(hit_reasons) if hit_reasons else ""

        all_rows.append(record)

        # ③ 对于对比类问题，添加对比记录
        if "问题分类" in row and row["问题分类"] == "客户归属组合漂移":
            # 检查是否有历史主_xxx 字段
            compare_record: dict[str, Any] = {"_record_type": "对比"}

            for col in original_cols:
                # 尝试从历史主字段获取
                hist_col = f"历史主_{col}"
                if hist_col in row and pd.notna(row[hist_col]):
                    compare_record[col] = row[hist_col]
                else:
                    compare_record[col] = ""

            # 在命中原因中标注
            compare_record["命中原因"] = "【对比记录】上期主映射数据"

            all_rows.append(compare_record)

    if not all_rows:
        return pd.DataFrame()

    result = pd.DataFrame(all_rows)

    # 调整列顺序：原始列 + 命中原因
    final_cols = original_cols + ["命中原因"]
    # 确保所有列都存在
    for col in final_cols:
        if col not in result.columns:
            result[col] = ""

    return result[final_cols]
