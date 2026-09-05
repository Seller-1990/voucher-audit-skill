# -*- coding: utf-8 -*-
"""为月度项目利润异常登记表的明细行生成"收入成本表区域截图"并锚定到附件列。

用 Excel COM 真实渲染：按账载客户筛选收入成本表、对命中单元格设置红色填充
（与登记人手工标记一致）、对区域 CopyPicture 后经 Chart 导出 PNG——
截图与在 Excel 里看到的样子完全一致（含筛选按钮、列样式）。

用法：
    python tools/gen_reg_screenshots.py \
        --reg "月度项目利润异常登记表-1.xlsx" \
        --report "<报告目录>/凭证审核报告_xxx.xlsx" \
        --workdir "D:\\分析Work\\...\\202608收入成本表xxx" \
        [--out-dir 附件截图] [--limit N] [--dry-run]

流程：
  1. 从报告「疑似数据错误清单」建立 (主体,客户,规则,金额)->源行号 映射
     （归属一致性/主体切换等无实际客户的规则，从「新增规则明细」按账载客户
     注册宽松键，命中后高亮目标月全部行）；
  2. 从登记表找出本轮工具追加的明细行（发现人=凭证审核工具 且描述含"明细"）；
  3. Excel COM：筛选该客户 -> 红色标记命中 -> 区域截图 PNG；
  4. 图片锚定到该行"附件"列，保存登记表（被占用时另存副本并提示）。
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage


# ---------------------------------------------------------------- 数据准备

def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def load_income_cost(workdir: Path):
    """openpyxl 读 考核表输出.xlsx 收入成本表：返回 (header, rows)；rows 含源行号。"""
    wb = openpyxl.load_workbook(workdir / "考核表输出.xlsx", read_only=True, data_only=True)
    ws = wb["收入成本表"]
    data = list(ws.iter_rows(values_only=True))
    header = [str(x) if x is not None else "" for x in data[0]]
    rows = []
    for i, r in enumerate(data[1:], start=2):
        if r[0] is None and r[1] is None:
            continue
        rows.append({"_src_row": i, "cells": r})
    return header, rows


def _customer_col_match(header: list[str], rows: list[dict], booked: str) -> list[dict]:
    """该客户（账载或实际）在源表中的数据行。"""
    bcol = header.index("账载客户") if "账载客户" in header else 4
    acol = header.index("实际客户") if "实际客户" in header else 5
    return [r for r in rows
            if booked in (str(r["cells"][bcol] or "").strip(),
                          str(r["cells"][acol] or "").strip())]


def month_rows_for_customer(header: list[str], rows: list[dict], booked: str, month: int) -> set[int]:
    """宽松键命中用：该客户目标月的全部源行号。"""
    mcol = header.index("月") if "月" in header else 1
    return {r["_src_row"] for r in _customer_col_match(header, rows, booked)
            if int(_num(r["cells"][mcol])) == month}


def all_rows_for_customer(header: list[str], rows: list[dict], booked: str) -> set[int]:
    """该客户（账载或实际）在源表中的全部行号（含表头下的所有月）。"""
    return {r["_src_row"] for r in _customer_col_match(header, rows, booked)}


def parse_src_rows(txt: str) -> list[int]:
    return [int(x) for x in re.findall(r"\d+", str(txt))]


def strip_biz_suffix(cust: str) -> str:
    """去掉待登记异常客户列的 (业务类型) 后缀。"""
    return re.sub(r"（[^）]*）\s*$", "", str(cust or "")).strip()


def build_hit_map(report: Path):
    """(主体, 实际客户, 规则名, 金额) -> 源行号集合。

    疑似数据错误清单中按账载客户分组的规则（归属一致性/主体切换）没有实际客户值，
    额外从「新增规则明细」按账载客户注册宽松键（值为空集合，命中后高亮目标月全部行）。
    """
    wb = openpyxl.load_workbook(report, read_only=True, data_only=True)
    ws = wb["疑似数据错误清单"]
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x) for x in rows[0]]
    j = {c: i for i, c in enumerate(h)}
    hit = defaultdict(set)
    for r in rows[1:]:
        if not any(v is not None for v in r):
            continue
        cust = strip_biz_suffix(r[j["实际客户"]])
        entity = str(r[j["主体账簿"]] or "")
        rules = str(r[j["命中规则"]] or "")
        amount = r[j["本月净额收入"]]
        srcs = parse_src_rows(r[j["源行号"]])
        for rule in rules.split("、"):
            rule = rule.strip()
            if not rule:
                continue
            hit[(entity, cust, rule, round(_num(amount), 2))] |= set(srcs)
            hit[(entity, cust, rule, None)] |= set(srcs)

    # 宽松键：无实际客户的规则改按账载客户注册。
    ws2 = wb["新增规则明细"]
    rows2 = list(ws2.iter_rows(values_only=True))
    h2 = [str(x) for x in rows2[0]]
    j2 = {c: i for i, c in enumerate(h2)}
    loose_rules = {"客户归属一致性", "客户归属一致性检查", "客户换了主体，映射要跟着改"}
    for r in rows2[1:]:
        if not any(v is not None for v in r):
            continue
        rule = str(r[j2.get("规则名称")] or "").strip()
        if rule not in loose_rules:
            continue
        cust = str(r[j2.get("账载客户")] or "").strip()
        if not cust:
            continue
        entity = str(r[j2.get("主体账簿")] or "").strip()
        hit[(entity, cust, rule, None)] |= set()
        hit[("", cust, rule, None)] |= set()
    return hit


# ---------------------------------------------------------------- Excel COM

XL_SCREEN = 1   # CopyPicture Appearance
XL_BITMAP = 2   # CopyPicture Format
HIT_RED = 0x5050FF  # BGR: 纯红 RGB(255,80,80)，与登记人手工标记的红色风格一致


def _escape_criteria(name: str) -> str:
    """AutoFilter 条件中的通配符转义。"""
    return str(name).replace("~", "~~").replace("*", "~*").replace("?", "~?")


class ExcelShot:
    """打开收入成本表，按客户筛选 + 标红命中 + 区域截图。"""

    def __init__(self, workdir: Path, month: int = 8):
        import win32com.client as win32
        self.app = win32.Dispatch("Excel.Application")
        # CopyPicture(xlScreen) 依赖真实渲染：Excel 必须可见，否则剪贴板是空白位图
        self.app.Visible = True
        self.app.DisplayAlerts = False
        self.path = str(workdir / "考核表输出.xlsx")
        self.wb = self.app.Workbooks.Open(self.path, ReadOnly=True)
        self.ws = self.wb.Worksheets("收入成本表")
        self.month = month

        # 表头与关键列号
        used = self.ws.UsedRange
        self.hdr_row = used.Row
        self.last_row = used.Row + used.Rows.Count - 1
        self.cols = {}
        for c in range(1, used.Columns.Count + 1):
            name = str(used.Cells(1, c).Value or "").strip()
            if name:
                self.cols[name] = c

        # 高亮列/行策略用的规则集合
        self.highlight_col_rules = {
            "客户归属一致性", "客户换了主体，映射要跟着改", "疑似客商改名",
            "工资好像挂错了客户（序时账）", "零头成本挂在别的部门",
        }
        self.base_cols = ["主体账簿", "月", "内外", "三级科目", "账载客户", "实际客户", "部门", "项目"]
        self.money_cols = [
            "全额收入", "成本合计", "工资", "社保", "公积金", "项目返费",
            "项目福利费", "项目其他费用", "第三方挂靠成本", "结算人次", "社保人数",
            "净额收入", "项目毛利润",
        ]
        self._chart = None

    def shot(self, booked: str, actual: str, rule: str, srcs: set[int], cust_max_row: int, out_png: Path) -> None:
        ws = self.ws
        booked_col = self.cols.get("账载客户", 5)
        actual_col = self.cols.get("实际客户", 6)

        # 筛选：账载客户 = 目标客户（含其实际客户同名情形由 OR 数组覆盖）
        if ws.AutoFilterMode:
            ws.AutoFilterMode = False
        crit = list(dict.fromkeys([_escape_criteria(booked), _escape_criteria(actual)]))
        ws.UsedRange.AutoFilter(Field=booked_col, Criteria1=crit, Operator=7)

        # 显示列：基础 8 列；金额类规则追加关键金额列（其余列隐藏，Copy 后临时表只剩显示列）
        show = {self.cols[c] for c in self.base_cols if c in self.cols}
        if rule not in self.highlight_col_rules:
            show |= {self.cols[c] for c in self.money_cols if c in self.cols}
        lo, hi = min(show), max(show)
        block = ws.Range(ws.Cells(1, lo), ws.Cells(1, hi)).EntireColumn
        block.Hidden = True
        for c in show:
            ws.Columns(c).Hidden = False

        tmp = None
        try:
            # 红色标记：命中源行（归属类只涂实际客户单元格；其余整行）
            for r in srcs:
                if rule in self.highlight_col_rules and actual_col:
                    ws.Cells(r, actual_col).Interior.Color = HIT_RED
                else:
                    ws.Range(ws.Cells(r, lo), ws.Cells(r, hi)).Interior.Color = HIT_RED

            # 只复制可见单元格（隐藏行/列不带入临时表）
            tmp = self.wb.Worksheets.Add()
            ws.AutoFilter.Range.SpecialCells(12).Copy(tmp.Cells(1, 1))  # 12 = xlCellTypeVisible
            self.app.CutCopyMode = False
            # 同步列宽与表头行高，保证截图比例与原表一致
            for c, sc in enumerate(sorted(show), start=1):
                tmp.Columns(c).ColumnWidth = ws.Columns(sc).ColumnWidth
            tmp.Rows(1).RowHeight = ws.Rows(self.hdr_row).RowHeight

            trng = tmp.UsedRange
            # 截图：CopyPicture(xlBitmap) -> 剪贴板 -> PIL 取图存 PNG（真实 Excel 渲染）
            trng.CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
            from PIL import ImageGrab
            im = None
            for _ in range(5):
                time.sleep(0.2)
                im = ImageGrab.grabclipboard()
                if im is not None:
                    break
            if im is None:
                raise RuntimeError(f"剪贴板未取到图像: {out_png.name}")
            out_png.parent.mkdir(parents=True, exist_ok=True)
            im.save(out_png, "PNG")
        finally:
            if tmp is not None:
                tmp.Delete()
            block.Hidden = False

    def _visible_rows_of_month(self, month_col: int) -> set[int]:
        out = set()
        rng = self.ws.Range(self.ws.Cells(self.hdr_row + 1, month_col),
                            self.ws.Cells(self.last_row, month_col))
        for i in range(1, rng.Rows.Count + 1):
            row = rng.Row + i - 1
            if self.ws.Rows(row).Hidden:
                continue
            if int(_num(rng.Cells(i, 1).Value)) == self.month:
                out.add(row)
        return out

    def close(self):
        try:
            if self._chart is not None:
                self._chart.Delete()
        except Exception:
            pass
        try:
            self.wb.Close(SaveChanges=False)
            self.app.Quit()
        except Exception:
            pass


# ---------------------------------------------------------------- 主流程

def main() -> int:
    ap = argparse.ArgumentParser(description="为登记表明细行生成收入成本表区域截图（Excel 真实渲染）并锚定到附件列")
    ap.add_argument("--reg", required=True)
    ap.add_argument("--report", required=True)
    ap.add_argument("--workdir", required=True)
    ap.add_argument("--out-dir", default=None, help="截图输出目录（默认：登记表同目录/附件截图）")
    ap.add_argument("--month", type=int, default=8)
    ap.add_argument("--dry-run", action="store_true", help="只生成图片，不写回登记表")
    ap.add_argument("--limit", type=int, default=0, help="只处理前 N 条明细（调试用）")
    args = ap.parse_args()

    reg_path = Path(args.reg)
    out_dir = Path(args.out_dir) if args.out_dir else reg_path.parent / "附件截图"

    hit_map = build_hit_map(Path(args.report))
    header, src_rows = load_income_cost(Path(args.workdir))

    wb = openpyxl.load_workbook(reg_path)
    ws = wb["异常登记表"]
    # 合并单元格映射：问题类型列（J）按组合并后，明细行需从区域左上角取规则名
    merged_top = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 10 and rng.max_col == 10:
            top_val = ws.cell(row=rng.min_row, column=10).value
            for r in range(rng.min_row, rng.max_row + 1):
                merged_top[r] = top_val
    targets = []
    for i in range(2, ws.max_row + 1):
        who = str(ws.cell(row=i, column=8).value or "")
        desc = str(ws.cell(row=i, column=5).value or "")
        if who != "凭证审核工具" or not desc:
            continue
        # 汇总行形如 "可能存在错误N：..."；明细行是同批次内其余行
        if not re.match(r"^可能存在错误\d+：", desc):
            targets.append(i)
    if args.limit:
        targets = targets[: args.limit]
    print(f"待附图明细行: {len(targets)}")

    def _norm(s: str) -> str:
        return str(s).replace("检查", "").strip()

    shots = []  # (row_idx, booked, rule, srcs, cust_max_row)
    miss = 0
    for i in targets:
        entity = str(ws.cell(row=i, column=3).value or "").strip()
        booked = strip_biz_suffix(ws.cell(row=i, column=4).value)
        rule = str(merged_top.get(i) or ws.cell(row=i, column=10).value or "").strip()
        amount = _num(ws.cell(row=i, column=7).value)

        srcs = hit_map.get((entity, booked, rule, round(amount, 2))) or set()
        if not srcs:
            srcs = hit_map.get((entity, booked, rule, None)) or set()
        if not srcs:
            cands = [v for k, v in hit_map.items()
                     if k[1] == booked and _norm(k[2]) == _norm(rule)]
            if cands:
                srcs = set().union(*cands)
        if not srcs and any(k[1] == booked and _norm(k[2]) == _norm(rule) for k in hit_map):
            # 宽松键（无实际客户的归属/主体类规则）：高亮该客户目标月全部行
            srcs = month_rows_for_customer(header, src_rows, booked, args.month)
        if not srcs:
            miss += 1
            print(f"  [miss] 行{i}: 未匹配 ({booked[:14]}.., {rule[:12]}..)")
            continue
        cust_rows = all_rows_for_customer(header, src_rows, booked)
        if not cust_rows:
            miss += 1
            print(f"  [miss] 行{i}: 源表无该客户行 {booked[:16]}")
            continue
        shots.append((i, booked, rule, srcs, max(cust_rows)))

    # Excel COM 批量截图
    shotter = ExcelShot(Path(args.workdir), month=args.month)
    ok = 0
    try:
        for n, (i, booked, rule, srcs, cust_max_row) in enumerate(shots, 1):
            actual = booked  # 筛选按账载客户；实际客户同名时由 xlFilterValues 数组覆盖
            fname = f"row{i:03d}_{re.sub(r'[\\\\/:*?\"<>|（）() ]', '_', booked)[:24]}_{rule[:14]}.png"
            png = out_dir / fname
            try:
                shotter.shot(booked, actual, rule, srcs, cust_max_row, png)
                ok += 1
            except Exception as e:
                miss += 1
                print(f"  [shot-fail] 行{i}: {str(e)[:80]}")
            if n % 30 == 0:
                print(f"  ... {n}/{len(shots)}")
    finally:
        shotter.close()
    print(f"截图完成 {ok}，失败/未匹配 {miss}")

    if args.dry_run:
        return 0

    # 锚定到登记表附件列（被占用时另存副本）
    wb = openpyxl.load_workbook(reg_path)
    ws = wb["异常登记表"]
    for i, booked, rule, _srcs, _cmr in shots:
        fname = f"row{i:03d}_{re.sub(r'[\\\\/:*?\"<>|（）() ]', '_', booked)[:24]}_{rule[:14]}.png"
        png = out_dir / fname
        if not png.exists():
            continue
        img = XLImage(str(png))
        scale = min(1.0, 960 / img.width)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, f"F{i}")
    try:
        wb.save(reg_path)
        print(f"已写回 {reg_path}")
    except PermissionError:
        alt = reg_path.with_name(reg_path.stem + "（附真实截图）.xlsx")
        wb.save(alt)
        print(f"[警告] 登记表被占用（请关闭 Excel 后重跑锚定），已另存: {alt}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
