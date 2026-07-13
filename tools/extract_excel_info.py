#!/usr/bin/env python
"""Excel 文件信息提取工具 - 从标题行提取关键信息"""

import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook


def extract_headers_info(file_path: Path) -> dict:
    """提取 Excel 文件的标题行信息"""
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    result = {
        "file": str(file_path),
        "sheets": [],
        "total_rows": 0,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(cell.value or "") for cell in ws[1] if cell.value is not None]

        # 统计行数（跳过标题行）
        rows_list = list(ws.iter_rows(values_only=True))
        row_count = len(rows_list) - 1 if len(rows_list) > 0 else 0

        result["sheets"].append({
            "name": sheet_name,
            "columns": headers,
            "column_count": len(headers),
            "row_count": row_count,
            "sample_values": [str(cell.value or "") for cell in ws[2]] if row_count > 0 else [],
        })

        result["total_rows"] += row_count

    wb.close()
    return result


def print_table(result: dict, show_sample: bool = False):
    """格式化输出信息"""
    print(f"\n📄 {Path(result['file']).name}")
    print(f"   Sheet数: {len(result['sheets'])} | 总行数: {result['total_rows']}")
    print()

    for sheet in result["sheets"]:
        print(f"   📑 {sheet['name']}")
        print(f"      列 ({sheet['column_count']}): {', '.join(sheet['columns'])}")
        print(f"      数据行: {sheet['row_count']}")

        if show_sample and sheet['sample_values']:
            print(f"      样例数据: {', '.join(str(v)[:50] for v in sheet['sample_values'])}")
        print()


def validate_columns(file_path: Path, columns: list[str]) -> bool:
    """验证 Excel 是否包含指定的列"""
    result = extract_headers_info(file_path)
    found_columns = set()

    for sheet in result["sheets"]:
        found_columns.update(sheet["columns"])

    missing = [col for col in columns if col not in found_columns]
    present = [col for col in columns if col in found_columns]

    print(f"\n📄 {file_path.name}")
    print(f"✅ 找到的列 ({len(present)}): {', '.join(present)}")
    if missing:
        print(f"❌ 缺失的列 ({len(missing)}): {', '.join(missing)}")

    return len(missing) == 0


def main():
    parser = argparse.ArgumentParser(description="提取 Excel 文件信息")
    parser.add_argument("file", help="Excel 文件路径")
    parser.add_argument("--show-sample", action="store_true", help="显示每列的样例数据")
    parser.add_argument("--validate-cols", nargs="+", metavar="COL", help="验证文件是否包含指定的列")

    args = parser.parse_args()
    file_path = Path(args.file).resolve()

    if args.validate_cols:
        valid = validate_columns(file_path, args.validate_cols)
        sys.exit(0 if valid else 1)
    else:
        result = extract_headers_info(file_path)
        print_table(result, show_sample=args.show_sample)


if __name__ == "__main__":
    main()