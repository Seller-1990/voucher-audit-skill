from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from voucher_audit.excel_annotation_com import write_source_annotations
from voucher_audit.source_annotation import QueryTableAnnotationPlan, RowAnnotation, SourceAnnotationBundle


pytestmark = pytest.mark.excel_com


@pytest.mark.skipif(sys.platform != "win32", reason="Excel COM 仅支持 Windows")
def test_real_excel_com_annotation_round_trip(tmp_path: Path) -> None:
    if os.environ.get("VOUCHER_AUDIT_EXCEL_INTEGRATION") != "1":
        pytest.skip("设置 VOUCHER_AUDIT_EXCEL_INTEGRATION=1 后运行真实 Excel COM 测试")

    workbook_path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "收入成本表"
    sheet.append(["主体账簿", "月", "净额收入"])
    sheet.append(["A", 3, 100])
    sheet.add_table(Table(displayName="AuditInput", ref="A1:C2"))
    workbook.save(workbook_path)

    plan = QueryTableAnnotationPlan(
        workbook_path=workbook_path,
        worksheet_name="收入成本表",
        table_name="AuditInput",
        gap_columns=1,
        headers=("凭证审核异常项", "凭证审核规则ID", "凭证审核命中原因"),
        row_annotations=(RowAnnotation(0, "异常", "TEST_RULE", "集成测试"),),
        cell_highlights=(),
        possible_highlight_columns=(),
    )

    result = write_source_annotations(SourceAnnotationBundle(plans=(plan,)))

    assert result.ok
    verified = load_workbook(workbook_path, data_only=False)
    values = list(verified["收入成本表"].values)
    assert "凭证审核异常项" in values[0]
    assert "异常" in values[1]
